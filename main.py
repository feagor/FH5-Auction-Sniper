import os
import sys
import time
import logging
import threading
import random
from logging.handlers import RotatingFileHandler
import configparser
from bisect import bisect_right

import cv2
import numpy as np
import pandas as pd
from openpyxl import load_workbook

import pygetwindow as gw
from mss import mss
import pydirectinput as pydi
import get_monitors as gm
import colorama

from overlay import OverlayController

class InputDriver:
    """Wraps keyboard/mouse automation with consistent timing helpers."""

    def __init__(self, keyboard, pointer):
        self.keyboard = keyboard
        self.pointer = pointer

    def wait(self, seconds: float) -> None:
        time.sleep(max(0.0, seconds))

    def tap(self, key: str, count: int = 1, interval: float = 0.1) -> None:
        for _ in range(max(0, int(count))):
            self.keyboard.press(key)
            if interval:
                self.wait(interval)

    def step(self, inc_key: str, dec_key: str, delta: int, interval: float = 0.1) -> None:
        if delta > 0:
            self.tap(inc_key, delta, interval)
        elif delta < 0:
            self.tap(dec_key, abs(delta), interval)

    def hold(self, key: str, duration: float = 5) -> None:
        self.keyboard.keyDown(key)
        self.wait(duration)
        self.keyboard.keyUp(key)

    def mouse_move(self, x: int, y: int, duration: float = 0.01) -> None:
        self.pointer.moveTo(x, y, duration=duration)

    def click(self) -> None:
        self.keyboard.mouseDown()
        self.wait(0.05)
        self.keyboard.mouseUp()

    def burst(self, count: int, gap: float = 0.01) -> None:
        for _ in range(max(0, int(count))):
            self.click()
            if gap:
                self.wait(gap)

if getattr(sys, 'frozen', False):
    CURRENT_DIR = os.path.dirname(sys.executable)
else:
    CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))

# Read configuration from settings.ini
config = configparser.ConfigParser()
config.read(os.path.join(CURRENT_DIR, 'settings.ini'))

def get_config_value(section, key, default=None, value_type=str):
    """
    Safely read value from config. Returns default on any error.
    value_type can be bool, int, float or callable for conversion.
    """
    try:
        value = config.get(section, key)
        if value_type == bool:
            return value.lower() in ['true', '1', 'yes']
        return value_type(value)
    except Exception:
        return default

# Get config data with defaults
LOCAL               = get_config_value('GENERAL', 'LOCAL', 'ENG')
EXCEL_FILENAME      = get_config_value('GENERAL', 'EXCEL_FILENAME', 'FH5_all_cars_info_v4.xlsx')
EXCEL_SHEET_NAME    = get_config_value('GENERAL', 'EXCEL_SHEET_NAME', 'all_cars_info')
LOCAL_MAKE_COL      = get_config_value('GENERAL', 'LOCAL_MAKE_COL', 'MAKE LOC (ENG)')
DEBUG_MODE          = get_config_value('GENERAL', 'DEBUG_MODE', False, bool)
WAIT_RESULT_TIME    = get_config_value('GENERAL', 'WAIT_RESULT_TIME', 1.0, float)
MAX_BUYOUT_PRICE    = get_config_value('GENERAL', 'MAX_BUYOUT_PRICE', 1000000, int)
SHUFFLE_CAR_LIST    = get_config_value('GENERAL', 'SHUFFLE_CAR_LIST', False, bool)
SNIPE_MIN_LIMIT     = get_config_value('GENERAL', 'SNIPE_MIN_LIMIT', 30, int)
# Constants (colorama codes kept for terminal coloring)
RED_CODE = '\033[1;31;40m'
GREEN_CODE = '\033[1;32;40m'
YELLOW_CODE = '\033[1;33;40m'
BLUE_CODE = '\033[1;34;40m'
CYAN_CODE = '\033[1;36;40m'
COLOR_END_CODE = '\033[0m'
#Other constants
SNIPE_SEC_LIMIT     = SNIPE_MIN_LIMIT * 60
GAME_TITLE          = 'Forza Horizon 5'
PAUSE_EVENT = threading.Event()
STOP_EVENT = threading.Event()

# Paths
EXCEL_PATH = os.path.join(CURRENT_DIR, EXCEL_FILENAME)
# Templates paths
IMAGE_PATH_SA   = os.path.join(CURRENT_DIR, 'images', LOCAL, 'SA.png')
IMAGE_PATH_CF   = os.path.join(CURRENT_DIR, 'images', LOCAL, 'CF.png')
IMAGE_PATH_AT   = os.path.join(CURRENT_DIR, 'images', LOCAL, 'AT.png')
IMAGE_PATH_BF   = os.path.join(CURRENT_DIR, 'images', LOCAL, 'BF.png')
IMAGE_PATH_PB   = os.path.join(CURRENT_DIR, 'images', LOCAL, 'PB.png')
IMAGE_PATH_BS   = os.path.join(CURRENT_DIR, 'images', LOCAL, 'BS.png')
IMAGE_PATH_NB   = os.path.join(CURRENT_DIR, 'images', LOCAL, 'NB.png')
IMAGE_PATH_VS   = os.path.join(CURRENT_DIR, 'images', LOCAL, 'VS.png')
IMAGE_PATH_AO   = os.path.join(CURRENT_DIR, 'images', LOCAL, 'AO.png')
IMAGE_PATH_HMG  = os.path.join(CURRENT_DIR, 'images', LOCAL, 'HMG.png')
IMAGE_PATH_HMMF = os.path.join(CURRENT_DIR, 'images', LOCAL, 'HMMF.png')
IMAGE_PATH_HMBS = os.path.join(CURRENT_DIR, 'images', LOCAL, 'HMBS.png')

# Region globals
REGION_HOME_TABS           = (0,0,0,0)
REGION_AUCTION_MAIN        = (0,0,0,0)
REGION_AUCTION_CAR_DESCR   = (0,0,0,0)
REGION_AUCTION_ACTION_MENU = (0,0,0,0)
REGION_AUCTION_RESULT      = (0,0,0,0)

# Screenshot matching parameters
THRESHOLD = 0.8
WIDTH_RATIO, HEIGHT_RATIO = 1, 1
EMPTY_CAR_INFO = {
    'Excel_index': -1,
    'Make_Name': '',
    'Make_Loc': [0,0],
    'Model_FName': '',
    'Model_SName': '',
    'Model_Loc': 0,
    'Buyout_num': 0,
    'Bought_num': 0
}

win_size = {'left': 0, 'top': 0, 'width': 0, 'height': 0}
first_run = True
miss_times = 1
start_time = time.time()
bought_in_session = 0
sct = mss()
sniping_car = EMPTY_CAR_INFO.copy()
snipe_secs_left = SNIPE_SEC_LIMIT


def exit_script():
    log_and_print('error', 'Script exits in 2 seconds!', RED_CODE)
    in_dr.wait(2)
    log_and_print('error', 'Script stops!', RED_CODE)
    STOP_EVENT.set()
    sys.exit(0)


def something_wrong():
    global miss_times
    HOME_TAB_META = {
        0: ('Buy & Sell',1.5),
        1: ('Garage',3.5),
        2: ('My Festival',5)
    }
    #try to open auction page again, if locates in home/festival menu
    home_page_result = get_best_match_img_array(
        [IMAGE_PATH_HMBS, IMAGE_PATH_HMG, IMAGE_PATH_HMMF], 
        REGION_HOME_TABS
    )
    if home_page_result:
        _, page_idx = home_page_result
        page_name, hold_seconds = HOME_TAB_META.get(page_idx, (f'index {page_idx}', 5))
        log_and_print('warning', f'Now located in Home/Festival menu ({page_name}). Try to open Auction House.', RED_CODE)
        in_dr.hold('a', hold_seconds)
        in_dr.tap('w')
        in_dr.tap('enter')
        in_dr.wait(1)
    else:
        log_and_print('warning', f'Fail to match anything. {miss_times}-th try to press ESC to see whether it works!', RED_CODE)
        active_game_window()
        in_dr.tap('esc')
        in_dr.wait(2)
        miss_times += 1
    
    if miss_times >= 10:
        log_and_print('error', 'Fail to detect anything, try to restart the script or game!', RED_CODE)
        exit_script()


class ColorFormatter(logging.Formatter):
    """Inject ANSI colors for console handler only."""
    def format(self, record):
        message = super().format(record)
        color = getattr(record, 'color', None)
        if color:
            return f"{color}{message}{COLOR_END_CODE}"
        else:
            return message
        

def setup_logging(debug_mode: bool):
    """
    Configure logging: file handler (DEBUG+) and console handler (INFO or DEBUG).
    Returns the configured logger.
    """
    log_path = os.path.join(CURRENT_DIR, 'fh5_sniper.log')
    logger = logging.getLogger('fh5_sniper')
    logger.setLevel(logging.DEBUG)  # capture all levels, handlers decide output
    # Remove existing handlers to avoid duplicate logs on re-import
    if logger.hasHandlers():
        logger.handlers.clear()
    # Rotating file handler: keep DEBUG-level history
    fh = RotatingFileHandler(log_path, maxBytes=5 * 1024 * 1024, backupCount=3, encoding='utf-8')
    fh.setLevel(logging.DEBUG)
    fh_formatter = logging.Formatter('%(asctime)s %(levelname)s %(name)s: %(message)s')
    fh.setFormatter(fh_formatter)
    logger.addHandler(fh)    
    # Console handler: INFO by default, DEBUG if requested
    ch = logging.StreamHandler()
    ch.setLevel(logging.DEBUG if debug_mode else logging.INFO)
    ch_formatter = ColorFormatter('%(asctime)s %(levelname)s: %(message)s', datefmt='%H:%M:%S')
    ch.setFormatter(ch_formatter)
    logger.addHandler(ch)
    return logger


def log_and_print(level: str, message: str, color: str = None):
    """
    Helper: log to logger while keeping optional coloring for console output only.
    level: 'debug', 'info', 'warning', 'error'
    color: one of color code constants or None
    """
    log_fn = {
        'debug': logger.debug,
        'info': logger.info,
        'warning': logger.warning,
        'error': logger.error,
    }.get(level, logger.info)    
    extra = {'color': color} if color else None
    if extra:
        log_fn(message, extra=extra)
    else:
        log_fn(message)


def capture_screen(region=None):
    if region:
        left, top, width, height = region
        monitor = {"left": left, "top": top, "width": width, "height": height}
    else:
        monitor = sct.monitors[0]
    shot = sct.grab(monitor)
    return cv2.cvtColor(np.array(shot), cv2.COLOR_BGRA2BGR)


def debug_screenshot(prefix_name, screenshot_cv):
    if DEBUG_MODE:
        debug_dir = os.path.join(CURRENT_DIR, 'debug', 'screen')
        os.makedirs(debug_dir, exist_ok=True)
        ts = time.strftime("%Y%m%d_%H%M%S")
        ms = int((time.time() % 1) * 1000)
        out_name = f"region_{prefix_name}_{ts}_{ms:03d}.png"
        out_path = os.path.join(debug_dir, out_name)
        # save BGR image
        cv2.imwrite(out_path, screenshot_cv)
    else: pass


def get_template_match(image_path, region=None, width_ratio=WIDTH_RATIO, height_ratio=HEIGHT_RATIO):
    """
    Take a screenshot of region, read template and run cv2.matchTemplate.
    Returns result matrix.
    """
    screenshot_cv = capture_screen(region=region)
    template = cv2.imread(image_path, cv2.IMREAD_COLOR)
    screenshot_cv = cv2.resize(screenshot_cv, (int(screenshot_cv.shape[1]/width_ratio), int(screenshot_cv.shape[0]/height_ratio)))
    result = cv2.matchTemplate(screenshot_cv, template, cv2.TM_CCOEFF_NORMED)
    base_name = image_path.rsplit('\\', 1)[-1].rsplit('.', 1)[0]
    debug_screenshot(base_name, screenshot_cv)
    min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(result)
    log_and_print('debug', f"Max match for {base_name}: {max_val*100:.1f} at {max_loc}")
    return result


def get_best_match_img_array(
    images_path,
    region=None,
    threshold=THRESHOLD
):
    """
    Find best match among one or multiple template images inside region.
    Returns location (x,y) or (location, index) when multiple images provided.
    """
    images_path = [images_path] if isinstance(images_path, str) else images_path
    return_index = True if len(images_path) > 1 else False
    best_prob,best_index,i = 0,0,0
    best_loc = ()
    for each_image_path in images_path:
        result = get_template_match(each_image_path, region=region)
        loc = np.where(result >= threshold)
        for pt in zip(*loc[::-1]):
            if best_prob < result[pt[1], pt[0]]:
                best_prob = result[pt[1], pt[0]]
                best_loc = (int(pt[0]), int(pt[1]))
                best_index = i
        i += 1
    if best_loc: 
        filename = images_path[0].rsplit('\\', 1)[-1]
        log_and_print('debug', f"Best match for { filename } at location { best_loc } with {best_prob*100:.1f}% probability.")
        if return_index:
            return best_loc, best_index
        return best_loc
    return None


def press_image(image_path, search_region):
    best_loc = get_best_match_img_array(image_path, search_region)
    #left, top, width, height = search_region
    if best_loc:
        in_dr.tap('enter')
        return True
    return False


def active_game_window():
    try:
        windows = gw.getWindowsWithTitle(GAME_TITLE)
        if not windows:
            logger.exception(f"Window {GAME_TITLE} not found")
            exit_script()
        game_window = windows[0]
        try:
            game_window.activate()
        except Exception:
            try:
                game_window.minimize()
                game_window.restore()                
            except Exception:
                logger.exception("Failed to activate/restore window")
                exit_script()
        win_size.update(
            {'left': game_window.left, 
            'top': game_window.top, 
            'width': game_window.width, 
            'height': game_window.height}
        )

        return game_window
    except Exception:
        logger.exception("Error getting game window")
        exit_script()


def measure_game_window():
    global REGION_HOME_TABS, REGION_AUCTION_MAIN, REGION_AUCTION_CAR_DESCR, REGION_AUCTION_ACTION_MENU, REGION_AUCTION_RESULT
    try:
        game_window = active_game_window()
        if game_window:
            game_window.resizeTo(1616, 939)
            win_size.update(
                {'left': game_window.left, 
                'top': game_window.top, 
                'width': game_window.width, 
                'height': game_window.height}
            )
            # Set regions based on measured window size    
            REGION_HOME_TABS = (
                520 + win_size['left'],
                164 + win_size['top'],
                570,
                40,
            )
            REGION_AUCTION_MAIN = (
                230 + win_size['left'],
                590 + win_size['top'],
                910,
                310,
            )
            REGION_AUCTION_CAR_DESCR = (
                790 + win_size['left'],
                190 + win_size['top'],
                810,
                90,
            )
            REGION_AUCTION_ACTION_MENU = (
                525 + win_size['left'],
                330 + win_size['top'],
                530,
                190,
            )
            REGION_AUCTION_RESULT = (
                60 + win_size['left'],
                150 + win_size['top'],
                180,
                40,
            )
            return game_window
        else:
            log_and_print('error', "Game window not found. Check the title.", RED_CODE)
            exit_script()
    except Exception:
        logger.exception("An error occurred while measuring the game window")
        exit_script()


def wait_if_paused(poll_interval: float = 0.1):
    """Block the automation loop while the pause overlay button is active."""
    while PAUSE_EVENT.is_set() and not STOP_EVENT.is_set():
        in_dr.wait(poll_interval)


def refresh_snipe_time_left():
    """Sync the script's countdown with the overlay-managed timer."""
    global snipe_secs_left
    controller = globals().get('overlay_controller')
    if controller:
        remaining = controller.get_remaining_seconds()
        if remaining is not None:
            snipe_secs_left = remaining
            return remaining
    elapsed = max(0, int(SNIPE_SEC_LIMIT - (time.time() - start_time)))
    snipe_secs_left = elapsed
    return elapsed


def pre_check():
    log_and_print('info', 'Welcome to the Forza 5 CAR BUYOUT Sniper', YELLOW_CODE)
    log_and_print('info', 'Running pre-check: monitor and game resolution', BLUE_CODE)
    monitors = gm.get_monitors()
    log_and_print('debug', f"Monitors data: {monitors}", CYAN_CODE)
    active_game_window()
    log_and_print('info', f"Game window {GAME_TITLE} found at ({win_size['left']},{win_size['top']}) with resolution {win_size['width']}x{win_size['height']}", CYAN_CODE)
    cx = win_size['left'] + win_size['width'] / 2
    cy = win_size['top'] + win_size['height'] / 2
    mon = gm.find_monitor_for_point(cx, cy, monitors)
    if mon:
        log_and_print('info', f"Center of the game window is at ({cx:.0f}, {cy:.0f}) on monitor: {mon.get('name')} {mon.get('resolution')}", CYAN_CODE)
    else:
        log_and_print('error', f"Center of the game window is at ({cx:.0f}, {cy:.0f}) not found on any monitor", RED_CODE)
        exit_script()
    measure_game_window()
    log_and_print('info', f"Resize {GAME_TITLE} resolution to {win_size['width']}x{win_size['height']} pixels!", CYAN_CODE)
    
    if overlay_controller.available:
        overlay_controller.launch((win_size['left'], win_size['top'], win_size['width'], win_size['height']))
    else:
        logger.debug('Overlay disabled: tkinter module not available')
    
    log_and_print('info', 'The script will start in 5 seconds', YELLOW_CODE)
    in_dr.wait(5)
    log_and_print('info', 'Script started', YELLOW_CODE)


def load_cars_from_excel():
    try:
        df = pd.read_excel(EXCEL_PATH, EXCEL_SHEET_NAME)
    except FileNotFoundError:
        log_and_print('error', f'Excel file not found at {EXCEL_PATH}', RED_CODE)
        exit_script()
    except ValueError as exc:
        log_and_print('error', f'Failed to load sheet "{EXCEL_SHEET_NAME}": {exc}', RED_CODE)
        exit_script()
    except Exception as exc:
        log_and_print('error', f'Failed to read Excel workbook: {exc}', RED_CODE)
        exit_script()

    required_columns = {
        'CAR MAKE',
        'CAR MODEL(Full Name)',
        'CAR MODEL(Short Name)',
        'MODEL LOC',
        'BUYOUT NUM',
        LOCAL_MAKE_COL,
    }
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        log_and_print('error', f'Missing required columns in Excel sheet: {", ".join(missing_columns)}', RED_CODE)
        exit_script()

    valid_rows = df[(df['BUYOUT NUM'] > 0) & (df['MODEL LOC'] != -1)]
    if valid_rows.empty:
        log_and_print('error', 'No cars found for sniping (check BUYOUT NUM and MODEL LOC).', RED_CODE)
        exit_script()

    cars = []
    for idx, row in valid_rows.iterrows():        
        car_info = EMPTY_CAR_INFO.copy()
        car_info['Excel_index'] = int(idx)
        car_info['Make_Name'] = row['CAR MAKE']
        make_loc = [int(part) for part in row[LOCAL_MAKE_COL].strip('()').split(',')]
        car_info['Make_Loc'] = make_loc
        car_info['Model_FName'] = row['CAR MODEL(Full Name)']
        car_info['Model_SName'] = row['CAR MODEL(Short Name)']
        car_info['Model_Loc'] = row['MODEL LOC']
        car_info['Buyout_num'] = int(row['BUYOUT NUM'] or 0)
        car_info['Bought_num'] = 0
        cars.append(car_info)

    if not cars:
        log_and_print('error', 'Car list is empty after processing the workbook.', RED_CODE)
        exit_script()

    if SHUFFLE_CAR_LIST:
        random.shuffle(cars)
        log_and_print('info', f'Car list shuffled (total {len(cars)} entries).', CYAN_CODE)

    return cars


def get_next_car_idx(cars, snipe_idx)->int:
    if not cars:
        return -1
    max_len = len(cars)
    start_idx = (snipe_idx + 1) % max_len
    idx = start_idx

    while True:
        if cars[idx]['Buyout_num'] > 0:
            return idx
        idx = (idx + 1) % max_len
        if idx == start_idx:
            break
    return -1


def set_auc_search_cond(new_car, old_car):
    global first_run, start_time
    is_confirm_button_found = get_best_match_img_array(IMAGE_PATH_CF, REGION_AUCTION_MAIN)
    
    if STOP_EVENT.is_set():
        return
    
    if is_confirm_button_found:
        # reset cursor
        active_game_window()
        in_dr.mouse_move(win_size['left'] + 10, win_size['top'] + 40)
        in_dr.burst(3)        
    else:
        something_wrong()  
    
    if first_run:
        log_and_print('info', 'Reseting search conditions', YELLOW_CODE)
        in_dr.tap('y', 1, 1) #reset search
    
    log_and_print('info', f'Setting search to: {new_car.get("Make_Name")}, {new_car.get("Model_FName")}', YELLOW_CODE)
    Make_X_Delta = int(old_car['Make_Loc'][0] - new_car['Make_Loc'][0])
    Make_Y_Delta = int(old_car['Make_Loc'][1] - new_car['Make_Loc'][1])
    in_dr.hold('w', 1.5) #goto make
    #not the same make
    if Make_X_Delta != 0 or Make_Y_Delta != 0: 
        in_dr.tap('enter')
        in_dr.wait(0.5)
        in_dr.step('w', 's', Make_Y_Delta)
        in_dr.wait(0.5)
        in_dr.step('a', 'd', Make_X_Delta)
        in_dr.wait(1)
        in_dr.tap('enter', 1, 0)
        in_dr.wait(0.5)
    
    if STOP_EVENT.is_set():
        return
    
    in_dr.tap('s') #goto model
    in_dr.wait(0.5)
    if Make_X_Delta == 0 and Make_Y_Delta == 0:
        Model_X_Delta = new_car['Model_Loc'] - old_car['Model_Loc']
    else:
        Model_X_Delta = new_car['Model_Loc']
    log_and_print('debug', f'New car model loc - {new_car['Model_Loc']}, Prev car model loc - {old_car['Model_Loc']}, Delta move - {Model_X_Delta}', CYAN_CODE)
    in_dr.step('d', 'a', Model_X_Delta, 0.2)
    
    if STOP_EVENT.is_set():
        return
    
    if first_run:
        in_dr.tap('s', 4, 0.3) #goto buyout price
        in_dr.wait(0.5)
        prices = [
            1000, 2000, 3000, 4000, 5000, 6000, 7000, 8000, 9000, 10000,
            11000, 20000, 30000, 40000, 50000, 60000, 70000, 80000, 90000, 100000,
            110000, 200000, 300000, 400000, 500000, 600000, 700000, 800000, 900000, 1000000,
            1100000, 2000000, 3000000, 4000000, 5000000, 6000000, 7000000, 8000000, 9000000, 10000000,
            11000000, 20000000, 30000000
        ]
        price_idx = bisect_right(prices, MAX_BUYOUT_PRICE) - 1
        if MAX_BUYOUT_PRICE != prices[price_idx]:
            log_and_print('info', f'Closest to parameter MAX_BUYOUT_PRICE = {MAX_BUYOUT_PRICE} game price is {prices[price_idx]}', YELLOW_CODE)
        log_and_print('info', f'Set buyout price to {prices[price_idx]}', YELLOW_CODE)            
        # setting buyout price, we have to tap one more time to set desire price
        # cause price_index starts from 0, not from 1
        in_dr.tap('d', price_idx+1, 0.15) 
        first_run = False
    
    in_dr.tap('s', 5, 0.3) #goto search button
    log_and_print('info', f'Start sniping {new_car.get("Make_Name")}, {new_car.get("Model_FName")}', GREEN_CODE)
    start_time = time.time()
    overlay_controller.update_status(
        car_name=new_car.get('Model_SName'),
        remaining_seconds=SNIPE_SEC_LIMIT,
        remaining_buyouts=new_car.get('Buyout_num'),
        purchased_count=new_car.get('Bought_num'),
        session_purchases=bought_in_session,
    )
    refresh_snipe_time_left()


def buyout(snipe_car) -> bool:
    def format_elapsed_time():
        seconds = time.time() - start_time
        return f'[{int(seconds // 60)}:{int(seconds % 60):02d}]'
    global bought_in_session
    result = False
    log_and_print('info', 'Car found in stock, trying to buyout...', GREEN_CODE)
    buyout_press_fl, buyout_res_fl = False, False
    iter = 0
    found_PB = found_VS = found_AO = None
    if STOP_EVENT.is_set():
        exit_script()
    
    # Press buyout button and wait result       
    while not buyout_press_fl and iter<=10:
        wait_if_paused()
        in_dr.wait(0.2)
        in_dr.tap('y')
        in_dr.wait(0.2)
        found_PB = get_best_match_img_array(IMAGE_PATH_PB, REGION_AUCTION_ACTION_MENU)
        found_VS = get_best_match_img_array(IMAGE_PATH_VS, REGION_AUCTION_ACTION_MENU)
        found_AO = get_best_match_img_array(IMAGE_PATH_AO, REGION_AUCTION_ACTION_MENU)
        if found_PB or found_VS or found_AO:
            in_dr.wait(0.3)
            # we are at active "place bid", need go down to buyout
            if found_PB:
                in_dr.tap('s')
                in_dr.tap('enter')
                in_dr.wait(0.5)
                in_dr.tap('enter')
                in_dr.wait(5)
                buyout_press_fl = True
            # we are at active View seller, this means someone bought before us
            elif found_VS:
                break
        iter += 1
   
    ## buyout button succesfully pressed, try to get buyout result
    if buyout_press_fl:
        iter = 0
        while not buyout_res_fl and iter<=10:
            wait_if_paused()
            iter += 1
            found_buyout_failed = get_best_match_img_array(IMAGE_PATH_BF, REGION_AUCTION_ACTION_MENU)
            found_buyout_success = get_best_match_img_array(IMAGE_PATH_BS, REGION_AUCTION_ACTION_MENU)
            if found_buyout_failed:                
                log_and_print('info', f'[{format_elapsed_time()}] BUYOUT Failed!', RED_CODE)
                in_dr.tap('enter')
                in_dr.wait(0.3)
                in_dr.tap('esc')
                buyout_res_fl = True
            if found_buyout_success:
                log_and_print('info', f'[{format_elapsed_time()}] BUYOUT Success!', GREEN_CODE)
                new_buyout_count = max(0, snipe_car['Buyout_num'] - 1)
                update_buyout(snipe_car['Excel_index'], new_buyout_count)
                snipe_car['Buyout_num'] = new_buyout_count
                snipe_car['Bought_num'] += 1
                bought_in_session += 1
                
                refresh_snipe_time_left()
                overlay_controller.update_status(
                    remaining_buyouts = new_buyout_count,
                    purchased_count   = snipe_car['Bought_num'],
                    session_purchases = bought_in_session,
                )
                buyout_res_fl = True
                result = True                   
                in_dr.tap('enter')
                in_dr.wait(0.3)
                in_dr.tap('esc')
                in_dr.wait(0.3)
                in_dr.tap('esc')
            else:
                log_and_print('info', f'[{format_elapsed_time()}] BUYOUT Missed!', YELLOW_CODE)
                in_dr.tap('esc')
                in_dr.wait(3)
    in_dr.wait(0.3)
    in_dr.tap('esc')
    return result


def update_buyout(row_index: int, buyout_num: int) -> None:
    try:
        wb = load_workbook(EXCEL_PATH)
        ws = wb[EXCEL_SHEET_NAME]
        target_row = row_index + 2  # +1 for header, +1 because Excel rows are 1-based
        buyout_col = None
        for cell in ws[1]:
            if str(cell.value).strip().upper() == 'BUYOUT NUM':
                buyout_col = cell.column
                break
        if buyout_col is None:
            log_and_print('error', 'Column BUYOUT NUM not found in Excel sheet', RED_CODE)
            return
        ws.cell(row=target_row, column=buyout_col, value=int(buyout_num))
        wb.save(EXCEL_PATH)
        log_and_print('debug', f'Updated BUYOUT NUM at row {row_index} to {buyout_num}', GREEN_CODE)
    except Exception as exc:
        log_and_print('error', f'Failed to update BUYOUT NUM: {exc}', RED_CODE)


def main():
    global bought_in_session
    pre_check()
    swap_car_fl = True
    snipe_idx = 0
    prev_car = EMPTY_CAR_INFO.copy()
    cars = load_cars_from_excel()
    snipe_car = cars[0] if cars else EMPTY_CAR_INFO.copy()
    formatted_cars = ' '.join(
        f'{idx}. {car["Make_Name"]}, {car["Model_SName"]} - {car["Buyout_num"]} pct\n'
        for idx, car in enumerate(cars, 1)
    )
    log_and_print('info', f'Today car list for sniping:\n {formatted_cars}', CYAN_CODE)

    while not STOP_EVENT.is_set():
        wait_if_paused()
        refresh_snipe_time_left()
        if snipe_secs_left <= 0 and not first_run:
            swap_car_fl = True
            log_and_print('info', f'Snipe for {snipe_car["Model_SName"]} timed out. Switching to next car.', YELLOW_CODE)
        in_dr.wait(0.35)
        wait_if_paused()
        is_search_auc_pressed = press_image(IMAGE_PATH_SA, REGION_AUCTION_MAIN)
        in_dr.wait(0.5)
        wait_if_paused()        
        if STOP_EVENT.is_set():
            something_wrong()
        
        if not is_search_auc_pressed:
            something_wrong()
            continue

        if swap_car_fl:
            prev_car = cars[snipe_idx] if not first_run else EMPTY_CAR_INFO.copy()            
            snipe_idx = get_next_car_idx(cars, snipe_idx) if not first_run else 0
            if snipe_idx == -1:
                log_and_print('info', f'All cars ({bought_in_session} pct) have been succesfully bought. Exit script.', GREEN_CODE)
                break
            snipe_car = cars[snipe_idx]
            set_auc_search_cond(snipe_car, prev_car)
            swap_car_fl = False
            if STOP_EVENT.is_set():
                break
        
        is_confirm_button_pressed = press_image(IMAGE_PATH_CF, REGION_AUCTION_MAIN)
        in_dr.wait(WAIT_RESULT_TIME)
        wait_if_paused()
        if STOP_EVENT.is_set():
            break
        
        is_auc_res_found = get_best_match_img_array(IMAGE_PATH_NB, REGION_AUCTION_RESULT)
        if is_auc_res_found:
            logger.debug('Auction results found')
            is_car_found = get_best_match_img_array(IMAGE_PATH_AT, REGION_AUCTION_CAR_DESCR)
            if is_car_found:
                buyout_succeeded = buyout(snipe_car)
                cars[snipe_idx] = snipe_car
                if buyout_succeeded and snipe_car['Buyout_num'] == 0:
                    swap_car_fl = True
                    log_and_print('info', f'Snipe for {snipe_car["Model_SName"]} Finished successfully. Switching to next car.', YELLOW_CODE)

            elif is_car_found is None and is_auc_res_found and is_confirm_button_pressed:
                log_and_print('debug', 'Car not found in stock')
                global miss_times
                miss_times = 1
                refresh_snipe_time_left()
                in_dr.tap('esc')
                in_dr.wait(0.5)
                continue
            overlay_controller.update_status(
                remaining_buyouts=snipe_car['Buyout_num'],
                purchased_count=snipe_car['Bought_num'],
                session_purchases=bought_in_session,
            )
        else:
            log_and_print('debug', 'Auction results not found :(')
            something_wrong()
            continue

    STOP_EVENT.set()
    log_and_print('info', 'Automation stopped.', YELLOW_CODE)


##INIT BLOCK##
logger = setup_logging(DEBUG_MODE)

overlay_controller = OverlayController(
    PAUSE_EVENT,
    STOP_EVENT,
    logger,
    log_callback=log_and_print,
    color_map={'resume': GREEN_CODE, 'pause': YELLOW_CODE, 'stop': RED_CODE},
    refocus_callback=active_game_window,
)
in_dr = InputDriver(pydi, pydi)
colorama.init(wrap=True)
pydi.PAUSE = 0

##END INIT BLOCK##

if __name__ == "__main__":
    main()
